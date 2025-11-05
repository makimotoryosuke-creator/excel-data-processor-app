#夕陽丘司法書士法人　エクセル単独加工アプリ
import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO

# --- 【カスタムCSSとページ設定】 ---

# ページの基本設定
st.set_page_config(
    page_title="売掛金入金データ作成ツール",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# カスタムCSSを適用して、フォント、ボタン、余白を調整
st.markdown("""
<style>
    /* 全体のフォントをモダンで読みやすいものに設定 */
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;700&display=swap');
    html, body, [class*="st-"] {
        font-family: 'Noto Sans JP', sans-serif;
    }

    /* メインコンテナの余白を調整し、中央に寄せる */
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        padding-left: 5rem;
        padding-right: 5rem;
    }

    /* タイトルとキャプションのスタイル */
    h1 {
        font-weight: 700;
        color: #2F3E46; /* ダークな青緑 */
        border-bottom: 2px solid #84A98C; /* アクセントカラー */
        padding-bottom: 0.5rem;
        margin-bottom: 0.5rem !important;
    }
    .stApp .stMarkdown p {
        color: #52796F; /* ダークな緑 */
        font-weight: 500;
    }

    /* ファイルアップローダーの強調表示 */
    .stFileUploader {
        border: 2px dashed #A7C9A0;
        padding: 2rem;
        border-radius: 8px;
        background-color: #F8F9FA;
        margin-top: 1.5rem;
        margin-bottom: 1.5rem;
    }

    /* ダウンロードボタンのスタイルを調整 */
    .stDownloadButton > button {
        background-color: #354F52; /* 濃いメインカラー */
        color: white;
        font-size: 1.1rem;
        font-weight: 600;
        padding: 0.8rem 2rem;
        border-radius: 8px;
        border: none;
        transition: background-color 0.2s;
    }
    .stDownloadButton > button:hover {
        background-color: #52796F; /* ホバー時の色 */
    }

    /* Statusメッセージ (Success/Error/Info) の見栄えを調整 */
    div[data-testid="stStatusContainer"] {
        font-size: 1rem;
        font-weight: 500;
        padding: 1rem;
        border-radius: 6px;
    }
</style>
""", unsafe_allow_html=True)


# --- 【B. Excel処理関数 - Webアプリ版 (ロジックは変更なし)】 ---

def process_excel_data(uploaded_file):
    """
    アップロードされたExcelファイルを処理し、加工後のExcelデータのバイナリを返す関数。
    """
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        
        # 日付計算（本日の前月）
        today = date.today()
        target_month_date = today - relativedelta(months=1)
        target_ym = target_month_date.strftime('%Y/%m')
        
        new_sheet_name = '読み込みシート'
        header_data = [
            "月日", "伝票番号", "証憑番号", "借方科目コード", "借方科目名", "借方補助コード", 
            "借方口座名", "借方部門コード", "借方部門名", "借方課税区分", "借方事業区分", 
            "借方消費税額自動計算か否か", "借方軽減税率か否か", "借方税率", "借方控除割合", 
            "借方取引金額", "借方消費税等", "借方税抜き金額", "貸方科目コード", "貸方科目名", 
            "貸方補助コード", "貸方口座名", "貸方部門コード", "貸方部門名", "貸方課税区分", 
            "貸方事業区分", "貸方消費税額自動計算か否か", "貸方軽減税率か否か", "貸方税率", 
            "貸方控除割合", "貸方取引金額", "貸方消費税等", "貸方税抜き金額", "取引先コード", 
            "取引先名", "取引先の事業者登録番号", "元帳摘要", "実際の仕入れ年月日表示区分", 
            "実際の仕入れ年月日１", "実際の仕入れ年月日２", "収支区分コード", "収支区分名", 
            "内訳区分コード", "内訳区分名"
        ] 

        ws_original = wb.worksheets[0]

        # --- O列基準のフィルタリングとデータ収集 ---
        rows_to_keep = []
        for row_index in range(2, ws_original.max_row + 1):
            o_cell = ws_original[f'O{row_index}'].value
            cell_ym = ''
            
            if o_cell:
                if isinstance(o_cell, date) or isinstance(o_cell, datetime):
                    cell_ym = o_cell.strftime('%Y/%m')
                elif isinstance(o_cell, str) and len(o_cell) >= 7:
                    cell_ym = o_cell[:7] 
            
            if cell_ym == target_ym:
                rows_to_keep.append(row_index)
                
        # --- 新しいブックとシートを作成し、データ転記 ---
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = new_sheet_name
        
        # ヘッダーを入力
        for col_index, header in enumerate(header_data, start=1):
            col_letter = get_column_letter(col_index)
            ws_new[f'{col_letter}1'] = header
            
        copy_list = [
            ('O', 'A'), ('P', 'P'), ('P', 'R'), ('P', 'AE'), ('P', 'AG'), ('B', 'AI'),
        ]
        
        new_row_index = 2
        for original_row_index in rows_to_keep:
            
            # --- データ転記 ---
            for original_col_letter, new_col_letter in copy_list:
                original_cell = ws_original[f'{original_col_letter}{original_row_index}']
                value_to_set = original_cell.value

                # AI列（取引先名）の文字数制限処理
                if new_col_letter == 'AI' and isinstance(value_to_set, str):
                    max_width = 32
                    current_width = 0
                    trimmed_value = ""
                    for char in value_to_set:
                        try:
                            width = len(char.encode('cp932', 'ignore')) 
                        except Exception:
                            width = 2
                        
                        if current_width + width <= max_width:
                            current_width += width
                            trimmed_value += char
                        else:
                            break
                    value_to_set = trimmed_value

                new_cell = ws_new[f'{new_col_letter}{new_row_index}']
                new_cell.value = value_to_set

            # --- 固定データ入力 ---
            seq_num = new_row_index - 1
            ws_new[f'B{new_row_index}'] = seq_num
            ws_new[f'C{new_row_index}'] = seq_num
            ws_new[f'D{new_row_index}'] = '1113' # 借方科目コード
            ws_new[f'E{new_row_index}'] = '普通預金' # 借方科目名
            ws_new[f'F{new_row_index}'] = '11' # 借方補助コード
            ws_new[f'G{new_row_index}'] = 'りそな銀行' # 借方口座名
            ws_new[f'J{new_row_index}'] = '0' # 借方課税区分
            ws_new[f'S{new_row_index}'] = '1122' # 貸方科目コード
            ws_new[f'T{new_row_index}'] = '売掛金' # 貸方科目名
            ws_new[f'Y{new_row_index}'] = '0' # 貸方課税区分
            ws_new[f'AK{new_row_index}'] = '売掛金入金' # 元帳摘要
            ws_new[f'AL{new_row_index}'] = '0' # 実際の仕入れ年月日表示区分
            
            new_row_index += 1

        # メモリ上にファイルを保存し、バイナリデータとして返す
        output = BytesIO()
        wb_new.save(output)
        output.seek(0)
        
        return output, new_row_index - 2, target_ym

    except Exception as e:
        # Streamlitのエラー表示機能を使用して、洗練された形でエラーを通知
        st.error(f"🚨 データ処理中に致命的なエラーが発生しました。ファイルの形式を確認してください。詳細: {e}")
        return None, 0, target_ym


# --- 【メインUIロジック】 ---

# タイトル
st.title("売掛金入金データ作成ツール 📊")
st.markdown("---")

# 説明
st.markdown("""
<p>
    このツールは、アップロードされた請求一覧ファイルから、**会計ソフトインポート用のデータ**を自動で作成します。<br>
    対象は、システム日付に基づき自動で決定された <span style='color: #2F3E46; font-weight: 700;'>前月分のデータ</span> のみです。
</p>
""", unsafe_allow_html=True)

# 処理対象年月の表示
target_month_for_display = date.today() - relativedelta(months=1)
st.info(f"✨ **現在の処理対象年月**: **`{target_month_for_display:%Y年%m月}`** のデータが抽出されます。", icon="📅")

# 1. ファイルアップロード
uploaded_file = st.file_uploader(
    "1. 処理したい Excel ファイル (.xlsx) をアップロードしてください",
    type="xlsx",
    accept_multiple_files=False # 複数ファイルは禁止
)

if uploaded_file is not None:
    st.markdown("### 2. 処理実行")
    # Streamlitのスピナー（ローディングアニメーション）を表示
    with st.spinner("🔄 データを処理し、会計ソフト用に整形しています..."):
        # 2. Excel処理の実行
        processed_data, processed_rows, target_ym = process_excel_data(uploaded_file)
    
    # 3. ダウンロードボタンの表示
    if processed_data is not None:
        st.success(f"✅ 処理が完了しました。**`{target_ym}`** のデータ **`{processed_rows}件`** を抽出しました。")
        
        # ダウンロードエリアの調整
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            label="⬇️ 処理済みファイルをダウンロード",
            data=processed_data,
            # ダウンロードファイル名をご要望に応じて変更
            file_name=f'会計インポート用_売掛金入金_{target_ym.replace("/", "_")}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        st.markdown("<p style='font-size:0.9rem; color: #777;'>※ ダウンロード後、保存先（デスクトップなど）をご確認ください。</p>", unsafe_allow_html=True)
        